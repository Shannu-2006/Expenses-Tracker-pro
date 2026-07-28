import io
import csv
import os
import re
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import threading
from datetime import datetime
from flask import (
    Flask,
    render_template,
    request,
    redirect,
    flash,
    session,
    send_file,
    Response,
    jsonify
)

from flask_sqlalchemy import SQLAlchemy

import pytesseract
from PIL import Image

# Configure Tesseract OCR binary path for Windows
tesseract_found = False
tesseract_paths = [
    r"C:\Program Files\Tesseract-OCR\tesseract.exe",
    r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
    os.path.join(os.environ.get("LOCALAPPDATA", ""), r"Programs\Tesseract-OCR\tesseract.exe"),
    os.path.join(os.environ.get("USERPROFILE", ""), r"AppData\Local\Programs\Tesseract-OCR\tesseract.exe"),
    r"C:\Users\bhuky\AppData\Local\Programs\Tesseract-OCR\tesseract.exe"
]
for path in tesseract_paths:
    if os.path.exists(path):
        pytesseract.pytesseract.tesseract_cmd = path
        tesseract_found = True
        break

from werkzeug.security import (
    generate_password_hash,
    check_password_hash
)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)
app.secret_key = "expense_tracker_secret"

app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///expenses.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)


class User(db.Model):
    id = db.Column(
        db.Integer,
        primary_key=True
    )

    username = db.Column(
        db.String(100),
        unique=True,
        nullable=False
    )

    email = db.Column(
        db.String(100),
        unique=True,
        nullable=False
    )

    password = db.Column(
        db.String(255),
        nullable=False
    )
class Budget(db.Model):
    id = db.Column(
        db.Integer,
        primary_key=True
    )

    amount = db.Column(
        db.Integer,
        nullable=False
    )

    month = db.Column(
        db.Integer,
        nullable=False
    )

    year = db.Column(
        db.Integer,
        nullable=False
    )

    user_id = db.Column(
        db.Integer,
        db.ForeignKey('user.id'),
        nullable=False
    )
class Expense(db.Model):
    id = db.Column(
        db.Integer,
        primary_key=True
    )

    name = db.Column(
        db.String(100),
        nullable=False
    )

    amount = db.Column(
        db.Integer,
        nullable=False
    )

    category = db.Column(
        db.String(50),
        nullable=False
    )

    date = db.Column(
        db.String(10),
        nullable=True
    )

    notes = db.Column(
        db.String(255),
        nullable=True
    )

    user_id = db.Column(
        db.Integer,
        db.ForeignKey('user.id'),
        nullable=False
    )

class Notification(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    message = db.Column(db.String(255), nullable=False)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)
    is_read = db.Column(db.Boolean, default=False)
    identifier = db.Column(db.String(50), nullable=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)

@app.route('/register', methods=['GET', 'POST'])
def register():

    if request.method == 'POST':

        username = request.form['username']
        email = request.form['email']
        password = request.form['password']
        confirm_password = request.form['confirm_password']

        if password != confirm_password:
            flash("Passwords do not match")
            return redirect('/register')

        existing_user = User.query.filter_by(
            email=email
        ).first()

        if existing_user:
            flash("Email already exists")
            return redirect('/register')
        existing_username = User.query.filter_by(
            username=username
        ).first()

        if existing_username:
            flash("Username already exists")
            return redirect('/register')

        hashed_password = generate_password_hash(
            password
        )

        user = User(
            username=username,
            email=email,
            password=hashed_password
        )

        db.session.add(user)
        db.session.commit()

        flash("Registration Successful")

        return redirect('/login')

    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():

    if request.method == 'POST':

        email = request.form['email']
        password = request.form['password']

        user = User.query.filter_by(
            email=email
        ).first()

        if user and check_password_hash(
            user.password,
            password
        ):

            session['user_id'] = user.id
            session['username'] = user.username

            return redirect('/')

        flash("Invalid Email or Password")

    return render_template('login.html')

@app.route('/logout')
def logout():

    session.clear()
@app.route('/delete_budget')
def delete_budget():
    if 'user_id' not in session:
        return redirect('/login')

    current_month = datetime.now().month
    current_year = datetime.now().year

    budget = Budget.query.filter_by(
        user_id=session['user_id'],
        month=current_month,
        year=current_year
    ).first()

    if budget:
        db.session.delete(budget)
        db.session.commit()
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({
                "success": True,
                "message": "Budget deleted successfully!",
                "stats": get_user_stats(session['user_id'])
            })
        flash("Budget deleted successfully!")
    else:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({
                "success": False,
                "message": "No budget found for this month.",
                "stats": get_user_stats(session['user_id'])
            })
        flash("No budget found for this month.")

    return redirect('/')

@app.route('/reset_budget')
def reset_budget():
    if 'user_id' not in session:
        return redirect('/login')

    current_month = datetime.now().month
    current_year = datetime.now().year

    budget = Budget.query.filter_by(
        user_id=session['user_id'],
        month=current_month,
        year=current_year
    ).first()

    if budget:
        budget.amount = 0
        db.session.commit()
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({
                "success": True,
                "message": "Budget reset successfully!",
                "stats": get_user_stats(session['user_id'])
            })
        flash("Budget reset successfully!")
    else:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({
                "success": False,
                "message": "No budget found for this month.",
                "stats": get_user_stats(session['user_id'])
            })
        flash("No budget found for this month.")

    return redirect('/')

@app.route('/set_budget', methods=['POST'])
def set_budget():
    if 'user_id' not in session:
        return redirect('/login')

    amount = int(request.form['budget'])

    current_month = datetime.now().month
    current_year = datetime.now().year

    budget = Budget.query.filter_by(
        user_id=session['user_id'],
        month=current_month,
        year=current_year
    ).first()

    if budget:
        budget.amount = amount
    else:
        budget = Budget(
            amount=amount,
            month=current_month,
            year=current_year,
            user_id=session['user_id']
        )
        db.session.add(budget)

    db.session.commit()

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({
            "success": True,
            "message": "Budget Saved Successfully!",
            "stats": get_user_stats(session['user_id'])
        })

    flash("Budget Saved Successfully!")
    return redirect('/')

# Helper function to compile full monthly analytics statistics in JSON
def get_user_stats(user_id):
    current_month = datetime.now().month
    current_year = datetime.now().year

    budget = Budget.query.filter_by(
        user_id=user_id,
        month=current_month,
        year=current_year
    ).first()

    budget_amount = 0
    if budget:
        budget_amount = budget.amount

    expenses = Expense.query.filter_by(
        user_id=user_id
    ).order_by(
        Expense.id.desc()
    ).all()

    total = sum(expense.amount for expense in expenses)
    transaction_count = len(expenses)
    highest_expense = max([expense.amount for expense in expenses], default=0)
    average_expense = total // transaction_count if transaction_count > 0 else 0

    category_totals = {
        "Food": 0,
        "Travel": 0,
        "Rent": 0,
        "Shopping": 0,
        "Fun": 0,
        "Other": 0
    }

    for expense in expenses:
        if expense.category in category_totals:
            category_totals[expense.category] += expense.amount

    budget_used = total
    remaining_budget = budget_amount - total
    if remaining_budget < 0:
        remaining_budget = 0

    actual_budget_percentage = 0
    budget_percentage = 0
    budget_status = "healthy"

    if budget_amount > 0:
        actual_budget_percentage = (budget_used / budget_amount) * 100
        budget_percentage = min(actual_budget_percentage, 100)
        if actual_budget_percentage >= 100:
            budget_status = "danger"
        elif actual_budget_percentage >= 80:
            budget_status = "warning"
        else:
            budget_status = "healthy"

    month_name = datetime.now().strftime("%B")
    if budget_amount > 0:
        budget_state = "Active"
    else:
        budget_state = "No Budget Set"

    expense_list = [{
        "id": e.id,
        "name": e.name,
        "amount": e.amount,
        "category": e.category,
        "date": e.date or '-',
        "notes": e.notes or ''
    } for e in expenses]

    return {
        "total": total,
        "transaction_count": transaction_count,
        "highest_expense": highest_expense,
        "average_expense": average_expense,
        "budget_amount": budget_amount,
        "budget_used": budget_used,
        "remaining_budget": remaining_budget,
        "budget_percentage": budget_percentage,
        "budget_status": budget_status,
        "actual_budget_percentage": actual_budget_percentage,
        "month_name": month_name,
        "current_year": current_year,
        "budget_state": budget_state,
        "category_totals": category_totals,
        "expenses": expense_list
    }

# Helper function to filter expenses for reports
def get_filtered_expenses_query(user_id, start_date, end_date, category):
    query = Expense.query.filter_by(user_id=user_id)
    if category and category != 'all':
        query = query.filter_by(category=category)
    if start_date:
        query = query.filter(Expense.date >= start_date)
    if end_date:
        query = query.filter(Expense.date <= end_date)
    return query.order_by(Expense.date.desc(), Expense.id.desc())

@app.route('/reports')
def reports():
    if 'user_id' not in session:
        return redirect('/login')
        
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    category = request.args.get('category', 'all')
    
    expenses = get_filtered_expenses_query(session['user_id'], start_date, end_date, category).all()
    
    total = sum(e.amount for e in expenses)
    count = len(expenses)
    highest = max((e.amount for e in expenses), default=0)
    average = total // count if count > 0 else 0
    
    category_totals = {cat: 0 for cat in ["Food", "Travel", "Rent", "Shopping", "Fun", "Other"]}
    for e in expenses:
        if e.category in category_totals:
            category_totals[e.category] += e.amount
            
    return render_template(
        'reports.html',
        expenses=expenses,
        total=total,
        count=count,
        highest_expense=highest,
        average_expense=average,
        category_totals=category_totals,
        start_date=start_date,
        end_date=end_date,
        category=category
    )

@app.route('/reports/export/csv')
def export_csv():
    if 'user_id' not in session:
        return redirect('/login')
        
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    category = request.args.get('category', 'all')
    
    expenses = get_filtered_expenses_query(session['user_id'], start_date, end_date, category).all()
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    writer.writerow(['Date', 'Expense Name', 'Category', 'Amount (INR)', 'Notes'])
    
    for e in expenses:
        writer.writerow([e.date or '-', e.name, e.category, e.amount, e.notes or ''])
        
    output.seek(0)
    
    filename = f"expense_report_{datetime.now().strftime('%Y%m%d')}.csv"
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-disposition": f"attachment; filename={filename}"}
    )

@app.route('/reports/export/excel')
def export_excel():
    if 'user_id' not in session:
        return redirect('/login')
        
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    category = request.args.get('category', 'all')
    
    expenses = get_filtered_expenses_query(session['user_id'], start_date, end_date, category).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Expense Report"
    
    # Premium Excel Styling
    font_title = Font(name='Segoe UI', size=16, bold=True, color='4F46E5')
    font_header = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    font_data = Font(name='Segoe UI', size=11)
    font_total = Font(name='Segoe UI', size=11, bold=True)
    
    fill_header = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    fill_summary = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
    
    border_thin = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )
    border_double_bottom = Border(
        top=Side(style='thin', color='9CA3AF'),
        bottom=Side(style='double', color='111827')
    )
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    
    # Title row
    ws.merge_cells('A1:E1')
    ws['A1'] = "EXPENSE TRACKER PRO - FINANCIAL STATEMENT"
    ws['A1'].font = font_title
    ws['A1'].alignment = align_left
    ws.row_dimensions[1].height = 30
    
    ws['A2'] = f"Generated On: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].font = Font(name='Segoe UI', size=9, italic=True, color='6B7280')
    ws.row_dimensions[2].height = 18
    
    # Selection summary
    ws['A4'] = "Start Date:"
    ws['A4'].font = font_total
    ws['B4'] = start_date if start_date else "All"
    ws['B4'].font = font_data
    
    ws['C4'] = "End Date:"
    ws['C4'].font = font_total
    ws['D4'] = end_date if end_date else "All"
    ws['D4'].font = font_data
    
    ws['A5'] = "Category:"
    ws['A5'].font = font_total
    ws['B5'] = category.capitalize()
    ws['B5'].font = font_data
    
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 20
    
    # Headers
    headers = ['Date', 'Expense Name', 'Category', 'Amount (INR)', 'Notes']
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center if col_idx in [1, 3] else (align_right if col_idx == 4 else align_left)
        cell.border = border_thin
    ws.row_dimensions[7].height = 25
    
    # Data Rows
    current_row = 8
    for e in expenses:
        ws.cell(row=current_row, column=1, value=e.date or '-').alignment = align_center
        ws.cell(row=current_row, column=2, value=e.name).alignment = align_left
        ws.cell(row=current_row, column=3, value=e.category).alignment = align_center
        
        amount_cell = ws.cell(row=current_row, column=4, value=e.amount)
        amount_cell.alignment = align_right
        amount_cell.number_format = '₹#,##0'
        
        ws.cell(row=current_row, column=5, value=e.notes or '-').alignment = align_left
        
        for col_idx in range(1, 6):
            c = ws.cell(row=current_row, column=col_idx)
            c.font = font_data
            c.border = border_thin
            
        ws.row_dimensions[current_row].height = 20
        current_row += 1
        
    # Total Spending Row
    ws.cell(row=current_row, column=3, value="Total Spending:").font = font_total
    ws.cell(row=current_row, column=3).alignment = align_right
    
    total_cell = ws.cell(row=current_row, column=4, value=f"=SUM(D8:D{current_row-1})")
    total_cell.font = font_total
    total_cell.alignment = align_right
    total_cell.border = border_double_bottom
    total_cell.number_format = '₹#,##0'
    
    ws.row_dimensions[current_row].height = 24
    
    # Adjust widths
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.row < 7:
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    filename = f"expense_report_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(
        file_stream,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )

@app.route('/reports/print')
def print_report():
    if 'user_id' not in session:
        return redirect('/login')
        
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    category = request.args.get('category', 'all')
    
    expenses = get_filtered_expenses_query(session['user_id'], start_date, end_date, category).all()
    total = sum(e.amount for e in expenses)
    
    current_month = datetime.now().month
    current_year = datetime.now().year
    budget = Budget.query.filter_by(
        user_id=session['user_id'],
        month=current_month,
        year=current_year
    ).first()
    budget_amount = budget.amount if budget else 0
    
    return render_template(
        'print_report.html',
        expenses=expenses,
        total=total,
        budget_amount=budget_amount,
        start_date=start_date,
        end_date=end_date,
        category=category,
        month_name=datetime.now().strftime("%B"),
        current_year=current_year
    )

@app.route('/analytics')
def analytics():
    if 'user_id' not in session:
        return redirect('/login')
        
    from datetime import timedelta
    today = datetime.now()
    current_year = today.year
    current_month = today.month
    
    expenses = Expense.query.filter_by(user_id=session['user_id']).all()
    
    # 1. Monthly Spending Trend (current year)
    monthly_trend = [0] * 12
    for e in expenses:
        if e.date:
            try:
                dt = datetime.strptime(e.date, '%Y-%m-%d')
                if dt.year == current_year:
                    monthly_trend[dt.month - 1] += e.amount
            except ValueError:
                pass
                
    months_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    months_full_names = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
    
    # 2. Weekly Spending Trend (Last 8 Weeks)
    weekly_trend_data = {}
    week_keys = []
    # Create keys for the last 8 weeks
    for i in range(7, -1, -1):
        start_of_week = today - timedelta(days=today.weekday() + i*7)
        key = start_of_week.strftime('%b %d')
        week_keys.append(key)
        weekly_trend_data[key] = 0
        
    for e in expenses:
        if e.date:
            try:
                dt = datetime.strptime(e.date, '%Y-%m-%d')
                for idx, k in enumerate(week_keys):
                    # Start date of this week slot
                    start_of_wk = today - timedelta(days=today.weekday() + (7 - idx)*7)
                    end_of_wk = start_of_wk + timedelta(days=6)
                    if start_of_wk.date() <= dt.date() <= end_of_wk.date():
                        weekly_trend_data[k] += e.amount
                        break
            except ValueError:
                pass
    weekly_trend_values = [weekly_trend_data[k] for k in week_keys]
    
    # 3. Yearly Spending Trend
    yearly_data = {}
    for e in expenses:
        if e.date:
            try:
                dt = datetime.strptime(e.date, '%Y-%m-%d')
                yr = str(dt.year)
                yearly_data[yr] = yearly_data.get(yr, 0) + e.amount
            except ValueError:
                pass
    sorted_years = sorted(yearly_data.keys())
    yearly_trend_values = [yearly_data[yr] for yr in sorted_years]
    
    # 4. Month-over-Month Comparison (This Month vs. Last Month category breakdown)
    prev_month = current_month - 1 if current_month > 1 else 12
    prev_month_year = current_year if current_month > 1 else current_year - 1
    
    categories = ["Food", "Travel", "Rent", "Shopping", "Fun", "Other"]
    current_month_cat = {cat: 0 for cat in categories}
    prev_month_cat = {cat: 0 for cat in categories}
    
    for e in expenses:
        if e.date:
            try:
                dt = datetime.strptime(e.date, '%Y-%m-%d')
                if dt.year == current_year and dt.month == current_month:
                    if e.category in current_month_cat:
                        current_month_cat[e.category] += e.amount
                elif dt.year == prev_month_year and dt.month == prev_month:
                    if e.category in prev_month_cat:
                        prev_month_cat[e.category] += e.amount
            except ValueError:
                pass
                
    current_month_values = [current_month_cat[c] for c in categories]
    prev_month_values = [prev_month_cat[c] for c in categories]
    
    # 5. Key Highlights Metrics
    current_month_total = sum(current_month_values)
    elapsed_days = today.day
    daily_average = current_month_total // elapsed_days if elapsed_days > 0 else 0
    
    highest_month_idx = monthly_trend.index(max(monthly_trend)) if sum(monthly_trend) > 0 else -1
    highest_month_name = months_full_names[highest_month_idx] if highest_month_idx != -1 else 'No Spends'
    highest_month_amount = monthly_trend[highest_month_idx] if highest_month_idx != -1 else 0
    
    active_months = [(idx, val) for idx, val in enumerate(monthly_trend) if val > 0]
    if active_months:
        lowest_month_idx, lowest_month_amount = min(active_months, key=lambda x: x[1])
        lowest_month_name = months_full_names[lowest_month_idx]
    else:
        lowest_month_name = 'No Spends'
        lowest_month_amount = 0
        
    annual_total = sum(monthly_trend)
    
    return render_template(
        'analytics.html',
        monthly_labels=months_names,
        monthly_values=monthly_trend,
        weekly_labels=week_keys,
        weekly_values=weekly_trend_values,
        yearly_labels=sorted_years,
        yearly_values=yearly_trend_values,
        categories=categories,
        current_month_values=current_month_values,
        prev_month_values=prev_month_values,
        daily_average=daily_average,
        annual_total=annual_total,
        highest_month_name=highest_month_name,
        highest_month_amount=highest_month_amount,
        lowest_month_name=lowest_month_name,
        lowest_month_amount=lowest_month_amount,
        current_month_name=today.strftime("%B"),
        current_year=current_year,
        prev_month_name=datetime(prev_month_year, prev_month, 1).strftime("%B")
    )

def parse_receipt_text(text):
    lines = [line.strip() for line in text.split('\n') if line.strip()]
    merchant_name = "Unknown Merchant"
    if lines:
        for line in lines[:5]:
            if len(line) > 3 and not re.match(r'^\d', line) and not any(kw in line.upper() for kw in ["TOTAL", "SUBTOTAL", "TAX", "ITEMS", "DATE", "TIME"]):
                merchant_name = line
                break
        if merchant_name == "Unknown Merchant":
            merchant_name = lines[0]
            
    amounts = []
    all_numbers = re.findall(r'₹?\s*(\d+[\.,]\d{2})\b', text)
    if not all_numbers:
        all_numbers = re.findall(r'\b(\d+)\b', text)
        
    for num_str in all_numbers:
        try:
            val = float(num_str.replace(',', ''))
            amounts.append(val)
        except ValueError:
            pass
            
    total_amount = 0
    total_match = None
    lines_upper = [l.upper() for l in lines]
    for idx, line in enumerate(lines_upper):
        if any(kw in line for kw in ["TOTAL", "GRAND TOTAL", "NET DUE", "TOTAL DUE", "AMOUNT DUE", "TOTAL AMOUNT", "PAID", "CASH DUE"]):
            nums_in_line = re.findall(r'(\d+[\.,]\d{2})\b', line)
            if not nums_in_line:
                nums_in_line = re.findall(r'\b(\d+)\b', line)
            if nums_in_line:
                try:
                    total_match = float(nums_in_line[-1].replace(',', ''))
                    break
                except ValueError:
                    pass
            if idx + 1 < len(lines_upper):
                next_line = lines_upper[idx + 1]
                nums_in_next = re.findall(r'(\d+[\.,]\d{2})\b', next_line)
                if not nums_in_next:
                    nums_in_next = re.findall(r'\b(\d+)\b', next_line)
                if nums_in_next:
                    try:
                        total_match = float(nums_in_next[0].replace(',', ''))
                        break
                    except ValueError:
                        pass
                        
    if total_match is not None:
        total_amount = int(total_match)
    elif amounts:
        filtered_amounts = [a for a in amounts if a < 100000]
        if filtered_amounts:
            total_amount = int(max(filtered_amounts))
            
    date_match = re.search(r'\b(\d{4})[-/](\d{2})[-/](\d{2})\b', text)
    date_str = ""
    if date_match:
        date_str = f"{date_match.group(1)}-{date_match.group(2)}-{date_match.group(3)}"
    else:
        date_match = re.search(r'\b(\d{2})[-/](\d{2})[-/](\d{4})\b', text)
        if date_match:
            date_str = f"{date_match.group(3)}-{date_match.group(2)}-{date_match.group(1)}"
            
    if not date_str:
        date_str = datetime.now().strftime('%Y-%m-%d')
        
    category = "Other"
    text_upper = text.upper()
    keywords = {
        "Food": ["FOOD", "CAFE", "COFFEE", "RESTAURANT", "BURGER", "PIZZA", "DINER", "BAKERY", "KITCHEN", "EATERY", "MCDONALD", "STARBUCKS", "GROCERY", "SUPERMARKET", "SWEETS"],
        "Travel": ["TRAVEL", "UBER", "OLA", "CAB", "TAXI", "METRO", "TRAIN", "BUS", "FLIGHT", "AIRLINE", "PETROL", "GAS", "FUEL", "TOLL", "DIESEL"],
        "Rent": ["RENT", "LEASE", "APARTMENT", "HOUSE", "MAINTENANCE", "PROPERTY"],
        "Shopping": ["SHOPPING", "STORE", "MALL", "CLOTHES", "APPAREL", "SHOES", "AMAZON", "FLIPKART", "ELECTRONICS", "GADGET", "MALL"],
        "Fun": ["FUN", "MOVIE", "CINEMA", "THEATRE", "CONCERT", "GAME", "PLAY", "BOWLING", "PARK", "EVENT", "CLUB", "BAR", "PUB"]
    }
    for cat, kws in keywords.items():
        if any(kw in text_upper for kw in kws):
            category = cat
            break
            
    return {
        "name": merchant_name[:50],
        "amount": total_amount,
        "category": category,
        "date": date_str,
        "notes": f"Scanned from receipt. Merchant: {merchant_name[:50]}"
    }

@app.route('/receipt/scan', methods=['POST'])
def scan_receipt():
    if 'user_id' not in session:
        return jsonify({"success": False, "message": "Unauthorized"}), 401
        
    try:
        pytesseract.get_tesseract_version()
    except Exception:
        return jsonify({
            "success": False,
            "error_type": "TESSERACT_NOT_FOUND",
            "message": "Tesseract OCR engine is not installed or not configured in your system path. Please install it to enable receipt scanning."
        }), 500

    if 'receipt' not in request.files:
        return jsonify({"success": False, "message": "No file uploaded"}), 400
        
    file = request.files['receipt']
    if file.filename == '':
        return jsonify({"success": False, "message": "Empty filename"}), 400
        
    temp_dir = os.path.join(os.path.dirname(__file__), "temp_uploads")
    os.makedirs(temp_dir, exist_ok=True)
    
    file_path = os.path.join(temp_dir, file.filename)
    file.save(file_path)
    
    try:
        img = Image.open(file_path)
        text = pytesseract.image_to_string(img)
        parsed_data = parse_receipt_text(text)
        os.remove(file_path)
        return jsonify({
            "success": True,
            "data": parsed_data
        })
    except Exception as e:
        if os.path.exists(file_path):
            os.remove(file_path)
        return jsonify({
            "success": False,
            "message": f"Error scanning receipt: {str(e)}"
        }), 500

# Asynchronous background email sender using python threading (Version 9)
def send_email_async(to_email, subject, body_html):
    def send_thread():
        smtp_server = os.environ.get("MAIL_SERVER")
        smtp_port = os.environ.get("MAIL_PORT", 587)
        smtp_user = os.environ.get("MAIL_USERNAME")
        smtp_password = os.environ.get("MAIL_PASSWORD")
        
        if not smtp_server or not smtp_user or not smtp_password:
            print("\n" + "="*60)
            print("📧 MOCK EMAIL NOTIFICATION DISPATCHED")
            print(f"Recipient : {to_email}")
            print(f"Subject   : {subject}")
            print(f"Body      :\n{body_html}")
            print("="*60 + "\n")
            return

        try:
            msg = MIMEMultipart('alternative')
            msg['Subject'] = subject
            msg['From'] = smtp_user
            msg['To'] = to_email

            part = MIMEText(body_html, 'html')
            msg.attach(part)

            server = smtplib.SMTP(smtp_server, int(smtp_port))
            server.starttls()
            server.login(smtp_user, smtp_password)
            server.sendmail(smtp_user, to_email, msg.as_string())
            server.quit()
            print(f"📧 Notification email successfully sent to {to_email}")
        except Exception as e:
            print(f"❌ Failed to dispatch email notification: {e}")

    threading.Thread(target=send_thread).start()

# Budget threshold warning and email alerts checker helper
def check_budget_thresholds(user_id, total_spent, budget_amount):
    if budget_amount <= 0:
        return
        
    current_month = datetime.now().month
    current_year = datetime.now().year
    percentage = (total_spent / budget_amount) * 100
    
    # 1. 100% Exceeded alert
    if percentage >= 100:
        tag_100 = f"budget_exceeded_100_{current_year}_{current_month}"
        exists = Notification.query.filter_by(user_id=user_id, identifier=tag_100).first()
        if not exists:
            readable_msg = f"🚨 Danger: Monthly budget limit exceeded! Spent ₹{total_spent} of ₹{budget_amount} ({percentage:.1f}%)."
            notif = Notification(message=readable_msg, user_id=user_id, identifier=tag_100)
            db.session.add(notif)
            db.session.commit()
            
            user = User.query.get(user_id)
            if user:
                subject = f"🚨 URGENT: Budget Limit Exceeded — ExpenseTracker Pro"
                body_html = f"""
                <div style="font-family: Arial, sans-serif; padding: 20px; color: #1e293b;">
                    <h2 style="color: #ef4444;">Budget Exceeded Alert</h2>
                    <p>Hello <strong>{user.username}</strong>,</p>
                    <p>This is an automated alert from your ExpenseTracker Pro ledger.</p>
                    <p style="background: #fef2f2; border: 1px solid #fee2e2; padding: 15px; border-radius: 8px; font-size: 16px;">
                        🚨 You have exceeded your monthly budget. You have spent <strong>₹{total_spent}</strong> of your monthly limit of <strong>₹{budget_amount}</strong>.
                    </p>
                    <p>Please review your expenses or adjust your budget targets to maintain financial control.</p>
                    <br>
                    <p>Regards,<br>ExpenseTracker Pro Team</p>
                </div>
                """
                send_email_async(user.email, subject, body_html)
                
    # 2. 80% Warning alert
    elif percentage >= 80:
        tag_80 = f"budget_warning_80_{current_year}_{current_month}"
        exists = Notification.query.filter_by(user_id=user_id, identifier=tag_80).first()
        if not exists:
            readable_msg = f"⚠️ Warning: Spent {percentage:.1f}% of your monthly budget (₹{total_spent} of ₹{budget_amount})."
            notif = Notification(message=readable_msg, user_id=user_id, identifier=tag_80)
            db.session.add(notif)
            db.session.commit()
            
            user = User.query.get(user_id)
            if user:
                subject = f"⚠️ WARNING: Budget Approaching Limit — ExpenseTracker Pro"
                body_html = f"""
                <div style="font-family: Arial, sans-serif; padding: 20px; color: #1e293b;">
                    <h2 style="color: #eab308;">Budget Threshold Warning</h2>
                    <p>Hello <strong>{user.username}</strong>,</p>
                    <p>This is an automated warning from your ExpenseTracker Pro ledger.</p>
                    <p style="background: #fefcbf; border: 1px solid #fef08a; padding: 15px; border-radius: 8px; font-size: 16px;">
                        ⚠️ You have spent <strong>{percentage:.1f}%</strong> of your monthly budget limit. You have spent <strong>₹{total_spent}</strong> of your set monthly limit of <strong>₹{budget_amount}</strong>.
                    </p>
                    <p>Try tracking your category totals to keep expenditures low.</p>
                    <br>
                    <p>Regards,<br>ExpenseTracker Pro Team</p>
                </div>
                """
                send_email_async(user.email, subject, body_html)

# API - Fetch recent notifications list
@app.route('/notifications')
def get_notifications():
    if 'user_id' not in session:
        return jsonify({"success": False, "message": "Unauthorized"}), 401
    notifs = Notification.query.filter_by(user_id=session['user_id']).order_by(Notification.id.desc()).limit(15).all()
    unread_count = Notification.query.filter_by(user_id=session['user_id'], is_read=False).count()
    
    notif_list = [{
        "id": n.id,
        "message": n.message,
        "timestamp": n.timestamp.strftime("%b %d, %H:%M"),
        "is_read": n.is_read
    } for n in notifs]
    
    return jsonify({
        "success": True,
        "notifications": notif_list,
        "unread_count": unread_count
    })

# API - Mark all notifications as read
@app.route('/notifications/read-all')
def read_all_notifications():
    if 'user_id' not in session:
        return jsonify({"success": False, "message": "Unauthorized"}), 401
    unread = Notification.query.filter_by(user_id=session['user_id'], is_read=False).all()
    for u in unread:
        u.is_read = True
    db.session.commit()
    return jsonify({"success": True, "message": "All notifications marked as read"})

# API - Trigger and dispatch monthly summary email statement
@app.route('/notifications/email-summary')
def email_summary():
    if 'user_id' not in session:
        return jsonify({"success": False, "message": "Unauthorized"}), 401
        
    user = User.query.get(session['user_id'])
    if not user:
        return jsonify({"success": False, "message": "User not found"}), 404
        
    stats = get_user_stats(session['user_id'])
    
    rows_html = ""
    for e in stats['expenses']:
        rows_html += f"""
        <tr>
            <td style="padding: 10px; border-bottom: 1px solid #eee;">{e['date']}</td>
            <td style="padding: 10px; border-bottom: 1px solid #eee;"><strong>{e['name']}</strong></td>
            <td style="padding: 10px; border-bottom: 1px solid #eee;"><span style="background: #e2e8f0; padding: 2px 6px; border-radius: 4px; font-size: 11px;">{e['category']}</span></td>
            <td style="padding: 10px; border-bottom: 1px solid #eee; text-align: right;">₹{e['amount']}</td>
            <td style="padding: 10px; border-bottom: 1px solid #eee;">{e['notes'] or '-'}</td>
        </tr>
        """
        
    subject = f"📊 Monthly Ledger Summary Statement — {stats['month_name']} {stats['current_year']}"
    body_html = f"""
    <div style="font-family: 'Segoe UI', Arial, sans-serif; padding: 25px; color: #1e293b; max-width: 600px; margin: 0 auto; border: 1px solid #e2e8f0; border-radius: 12px;">
        <h2 style="color: #4f46e5; border-bottom: 2px solid #4f46e5; padding-bottom: 10px; margin-top: 0;">ExpenseTracker Pro Statement</h2>
        <p>Hello <strong>{user.username}</strong>,</p>
        <p>Here is your financial ledger report summary compile for the month of <strong>{stats['month_name']} {stats['current_year']}</strong>.</p>
        
        <div style="display: grid; grid-template-columns: repeat(3, 1fr); gap: 10px; margin: 20px 0;">
            <div style="background: #f8fafc; border: 1px solid #e2e8f0; padding: 12px; border-radius: 8px; text-align: center; display: inline-block; width: 30%;">
                <span style="font-size: 11px; text-transform: uppercase; color: #64748b; font-weight: 600;">Total Spent</span>
                <div style="font-size: 18px; font-weight: bold; color: #ef4444; margin-top: 4px;">₹{stats['total']}</div>
            </div>
            <div style="background: #f8fafc; border: 1px solid #e2e8f0; padding: 12px; border-radius: 8px; text-align: center; display: inline-block; width: 30%;">
                <span style="font-size: 11px; text-transform: uppercase; color: #64748b; font-weight: 600;">Monthly Budget</span>
                <div style="font-size: 18px; font-weight: bold; color: #4f46e5; margin-top: 4px;">₹{stats['budget_amount']}</div>
            </div>
            <div style="background: #f8fafc; border: 1px solid #e2e8f0; padding: 12px; border-radius: 8px; text-align: center; display: inline-block; width: 30%;">
                <span style="font-size: 11px; text-transform: uppercase; color: #64748b; font-weight: 600;">Remaining</span>
                <div style="font-size: 18px; font-weight: bold; color: #10b981; margin-top: 4px;">₹{stats['remaining_budget']}</div>
            </div>
        </div>
        
        <h3>Compiled Ledger Items</h3>
        <table style="width: 100%; border-collapse: collapse; font-size: 13px;">
            <thead>
                <tr style="background: #f1f5f9; text-align: left;">
                    <th style="padding: 10px; font-weight: 600;">Date</th>
                    <th style="padding: 10px; font-weight: 600;">Name</th>
                    <th style="padding: 10px; font-weight: 600;">Category</th>
                    <th style="padding: 10px; font-weight: 600; text-align: right;">Amount</th>
                    <th style="padding: 10px; font-weight: 600;">Notes</th>
                </tr>
            </thead>
            <tbody>
                {rows_html if rows_html else '<tr><td colspan="5" style="padding: 20px; text-align: center; color: #94a3b8;">No expenses recorded this month.</td></tr>'}
            </tbody>
        </table>
        
        <br>
        <p style="font-size: 11px; color: #94a3b8; border-top: 1px solid #e2e8f0; padding-top: 15px; margin-top: 25px;">
            This statement was generated and dispatched automatically via ExpenseTracker Pro.
        </p>
    </div>
    """
    
    send_email_async(user.email, subject, body_html)
    
    tag_summary = f"statement_email_{datetime.now().strftime('%Y_%m_%d_%H_%M_%S')}"
    readable_msg = f"📧 Statement Dispatched: Emailed summary for {stats['month_name']} {stats['current_year']} to {user.email}."
    notif = Notification(message=readable_msg, user_id=session['user_id'], identifier=tag_summary)
    db.session.add(notif)
    db.session.commit()
    
    return jsonify({"success": True, "message": f"Summary statement email dispatched to {user.email}!"})

@app.route('/', methods=['GET', 'POST'])
def home():

    if 'user_id' not in session:
        return redirect('/login')

    if request.method == 'POST':

        name = request.form['name']
        amount = request.form['amount']
        category = request.form['category']
        date = request.form.get('date')
        notes = request.form.get('notes')

        if not date:
            date = datetime.now().strftime('%Y-%m-%d')

        expense = Expense(
            name=name,
            amount=int(amount),
            category=category,
            date=date,
            notes=notes,
            user_id=session['user_id']
        )

        db.session.add(expense)
        db.session.commit()

        # Check budget thresholds warnings (Version 9)
        try:
            current_month = datetime.now().month
            current_year = datetime.now().year
            budget = Budget.query.filter_by(user_id=session['user_id'], month=current_month, year=current_year).first()
            budget_amount = budget.amount if budget else 0
            
            # Recalculate current month total spent
            expenses = Expense.query.filter_by(user_id=session['user_id']).all()
            current_month_total = 0
            for e in expenses:
                if e.date:
                    try:
                        dt = datetime.strptime(e.date, '%Y-%m-%d')
                        if dt.year == current_year and dt.month == current_month:
                            current_month_total += e.amount
                    except ValueError:
                        pass
            check_budget_thresholds(session['user_id'], current_month_total, budget_amount)
        except Exception as e_thresh:
            print("Threshold warning check error:", e_thresh)

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({
                "success": True,
                "message": "Expense added successfully!",
                "stats": get_user_stats(session['user_id'])
            })

        flash("Expense added successfully!")

        return redirect('/')

    # -----------------------------
    # Get Current Month Budget
    # -----------------------------

    current_month = datetime.now().month
    current_year = datetime.now().year

    budget = Budget.query.filter_by(
        user_id=session['user_id'],
        month=current_month,
        year=current_year
    ).first()

    budget_amount = 0

    if budget:
        budget_amount = budget.amount

    # -----------------------------
    # Get Expenses
    # -----------------------------

    expenses = Expense.query.filter_by(
        user_id=session['user_id']
    ).order_by(
        Expense.id.desc()
    ).all()

    total = sum(
        expense.amount
        for expense in expenses
    )

    transaction_count = len(expenses)

    highest_expense = max(
        [expense.amount for expense in expenses],
        default=0
    )

    average_expense = (
        total // transaction_count
        if transaction_count > 0
        else 0
    )

    category_totals = {
        "Food": 0,
        "Travel": 0,
        "Rent": 0,
        "Shopping": 0,
        "Fun": 0,
        "Other": 0
    }

    for expense in expenses:

        if expense.category in category_totals:
            category_totals[
                expense.category
            ] += expense.amount

    highest_category = max(
        category_totals,
        key=category_totals.get
    )

    highest_category_amount = (
        category_totals[highest_category]
    )

    monthly_summary = {
        "total": total,
        "transactions": transaction_count,
        "highest_expense": highest_expense,
        "average_expense": average_expense
        
    }
    budget_used = total

    remaining_budget = budget_amount - total

    if remaining_budget < 0:
        remaining_budget = 0

       # -----------------------------
    # Budget Progress Percentage
    # -----------------------------

    actual_budget_percentage = 0
    budget_percentage = 0
    budget_status = "healthy"

    if budget_amount > 0:

        actual_budget_percentage = (
            budget_used / budget_amount
        ) * 100

        budget_percentage = min(
            actual_budget_percentage,
            100
        )

        if actual_budget_percentage >= 100:
            budget_status = "danger"

        elif actual_budget_percentage >= 80:
            budget_status = "warning"

        else:
            budget_status = "healthy"

    # -----------------------------
    # Current Budget Information
    # -----------------------------

    month_name = datetime.now().strftime("%B")
    current_year = datetime.now().year

    if budget_amount > 0:
        budget_state = "Active"
    else:
        budget_state = "No Budget Set"

    return render_template(
        'index.html',
        expenses=expenses,
        total=total,
        transaction_count=transaction_count,
        highest_expense=highest_expense,
        average_expense=average_expense,
        category_totals=category_totals,
        highest_category=highest_category,
        highest_category_amount=highest_category_amount,
        monthly_summary=monthly_summary,
        budget_amount=budget_amount,
        budget_used=budget_used,
        remaining_budget=remaining_budget,
        budget_percentage=budget_percentage,
        budget_status=budget_status,
        actual_budget_percentage=actual_budget_percentage,
        month_name=month_name,
        current_year=current_year,
        budget_state=budget_state
    )
@app.route('/update/<int:id>', methods=['POST'])
def update(id):

    if 'user_id' not in session:
        return redirect('/login')

    expense = Expense.query.filter_by(
        id=id,
        user_id=session['user_id']
    ).first_or_404()

    expense.name = request.form['name']
    expense.amount = int(
        request.form['amount']
    )
    expense.category = request.form['category']
    
    date = request.form.get('date')
    notes = request.form.get('notes')
    
    if date:
        expense.date = date
    expense.notes = notes

    db.session.commit()

    # Check budget thresholds warnings (Version 9)
    try:
        current_month = datetime.now().month
        current_year = datetime.now().year
        budget = Budget.query.filter_by(user_id=session['user_id'], month=current_month, year=current_year).first()
        budget_amount = budget.amount if budget else 0
        
        expenses = Expense.query.filter_by(user_id=session['user_id']).all()
        current_month_total = 0
        for e in expenses:
            if e.date:
                try:
                    dt = datetime.strptime(e.date, '%Y-%m-%d')
                    if dt.year == current_year and dt.month == current_month:
                        current_month_total += e.amount
                except ValueError:
                    pass
        check_budget_thresholds(session['user_id'], current_month_total, budget_amount)
    except Exception as e_thresh:
        print("Threshold warning check error on update:", e_thresh)

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({
            "success": True,
            "message": "Expense updated successfully!",
            "stats": get_user_stats(session['user_id'])
        })

    flash("Expense updated successfully!")

    return redirect('/')


@app.route('/delete/<int:id>')
def delete(id):

    if 'user_id' not in session:
        return redirect('/login')

    expense = Expense.query.filter_by(
        id=id,
        user_id=session['user_id']
    ).first_or_404()

    db.session.delete(expense)
    db.session.commit()

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({
            "success": True,
            "message": "Expense deleted successfully!",
            "stats": get_user_stats(session['user_id'])
        })

    flash("Expense deleted successfully!")

    return redirect('/')


if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        try:
            inspector = db.inspect(db.engine)
            columns = [c['name'] for c in inspector.get_columns('expense')]
            if 'date' not in columns:
                db.session.execute(db.text("ALTER TABLE expense ADD COLUMN date TEXT"))
            if 'notes' not in columns:
                db.session.execute(db.text("ALTER TABLE expense ADD COLUMN notes TEXT"))
            db.session.commit()
        except Exception as e:
            print("Auto-migration result:", e)
            db.session.rollback()
    app.run(debug=True)